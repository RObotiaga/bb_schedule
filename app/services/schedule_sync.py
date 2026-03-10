import asyncio
import os
import re
import shutil
import logging
import sqlite3
import ssl
from datetime import datetime, timezone
from urllib.parse import urljoin, unquote, quote

import aiohttp
from bs4 import BeautifulSoup

from app.core.config import DOWNLOAD_DIR, DB_PATH, BB_LOGIN, BB_PASSWORD, BB_URL
from app.core.logger import setup_logging
from app.core.repositories.job_log import save_job_log, cleanup_old_job_logs


logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
# --- CONSTANTS ---
MONTHS_MAP = {
    'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4, 'мая': 5, 'июня': 6,
    'июля': 7, 'августа': 8, 'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12
}

# Базовый путь CMS к расписанию очной формы
CMS_SCHEDULE_BASE = "/webapps/cmsmain/webui/institution/Расписание/Очная форма обучения"


class ScheduleFetcher:
    """
    Скачивает файлы расписания с Blackboard через HTTP (aiohttp).
    Заменяет Playwright — быстрее в ~10 раз, не создаёт процесс Chromium.
    """

    def __init__(self):
        self.download_dir = DOWNLOAD_DIR
        self.login = BB_LOGIN
        self.password = BB_PASSWORD
        self.base_url = BB_URL.rstrip("/")
        self._session: aiohttp.ClientSession | None = None

    def ensure_download_dir(self):
        if os.path.exists(self.download_dir):
            logging.info(f"Удаляем старую папку с расписаниями: {self.download_dir}")
            shutil.rmtree(self.download_dir)
        logging.info(f"Создаем пустую папку для расписаний: {self.download_dir}")
        os.makedirs(self.download_dir, exist_ok=True)

    async def _get_page(self, url: str) -> BeautifulSoup:
        """GET запрос и парсинг HTML."""
        async with self._session.get(url) as resp:
            text = await resp.text()
            return BeautifulSoup(text, "html.parser")

    async def login_to_bb(self):
        """Авторизация через POST-форму с nonce."""
        soup = await self._get_page(self.base_url)
        nonce_tag = soup.find("input", {"name": "blackboard.platform.security.NonceUtil.nonce"})
        nonce = nonce_tag["value"] if nonce_tag else ""

        login_data = {
            "user_id": self.login,
            "password": self.password,
            "login": "Войти",
            "action": "login",
            "new_loc": "",
            "blackboard.platform.security.NonceUtil.nonce": nonce,
        }
        async with self._session.post(
            f"{self.base_url}/webapps/login/", data=login_data, allow_redirects=True
        ) as resp:
            text = await resp.text()
            if "logout" not in text.lower() and "logoutLink" not in text:
                raise RuntimeError("BB login failed — не найден маркер авторизации")
        logging.info("BB login через HTTP — успешно")

    def _is_session_relevant(self, week_name: str) -> bool:
        """Фильтрует неактуальные сессии по семестру и году."""
        name_lower = week_name.lower()
        if "аттестация" not in name_lower and "сессия" not in name_lower:
            return True

        now = datetime.now()
        month, year = now.month, now.year

        if 9 <= month <= 12 or month == 1:
            target_semester = 1
            start_year = year - 1 if month == 1 else year
        else:
            target_semester = 2
            start_year = year - 1 if month < 9 else year

        match = re.search(r"(\d)\s*семестр.*?(\d{4})[/-](\d{4})", week_name, re.IGNORECASE)
        if match:
            sem = int(match.group(1))
            y_start = int(match.group(2))
            if sem == target_semester and y_start == start_year:
                return True
            logging.info(f"Пропуск неактуальной сессии: {week_name}")
            return False
        return False

    async def _get_week_folders(self) -> dict[str, str]:
        """
        Возвращает {week_name: cms_path} для актуальных недель.
        Пример: {"Нечетная неделя": ".../Нечетная неделя", "Четная неделя": "..."}
        """
        url = f"{self.base_url}{CMS_SCHEDULE_BASE}/?action=frameset&subaction=view"
        soup = await self._get_page(url)

        week_folders = {}
        # Ищем frameset-ссылки, которые ведут внутрь текущей папки
        for link in soup.select('a[href*="action=frameset"]'):
            href = unquote(link.get("href", ""))
            name = link.get_text(strip=True)
            if not name or not href:
                continue

            # Берём только вложенные папки (Нечетная/Четная/Аттестация)
            name_lower = name.lower()
            is_week = any(kw in name_lower for kw in ["неделя", "четная", "нечет", "аттестация", "сессия"])
            # Проверяем: после "Очная форма обучения/" есть ещё контент (имя подпапки)
            clean_href = href.split("?")[0]
            parent_marker = "Очная форма обучения/"
            is_subfolder = parent_marker in clean_href and clean_href.split(parent_marker, 1)[1].strip("/") != ""

            if is_week and is_subfolder and self._is_session_relevant(name):
                # Извлекаем путь до ?action=
                cms_path = href.split("?")[0]
                week_folders[name] = cms_path

        logging.info(f"Найдено актуальных недель: {len(week_folders)}")
        return week_folders

    async def _get_faculty_folders(self, week_cms_path: str) -> dict[str, str]:
        """
        Возвращает {faculty_name: cms_path} для факультетов внутри недели.
        """
        url = f"{self.base_url}{week_cms_path}?action=frameset&subaction=view"
        soup = await self._get_page(url)

        faculties = {}
        for link in soup.select('a[href*="action=frameset"]'):
            href = unquote(link.get("href", ""))
            name = link.get_text(strip=True)
            if not name or not href:
                continue
                
            if name.lower() in ("открыть", "open"):
                continue

            # Факультеты — подпапки текущей недели
            week_basename = week_cms_path.rstrip("/").split("/")[-1]
            if week_basename + "/" in href:
                subfolder = href.split(week_basename + "/")[-1].split("?")[0].rstrip("/")
                if subfolder and "/" not in subfolder:
                    faculties[name] = href.split("?")[0]

        return faculties

    async def _get_xls_links(self, folder_cms_path: str) -> list[tuple[str, str]]:
        """
        Возвращает [(filename, download_url)] для .xls/.xlsx внутри папки.
        """
        url = f"{self.base_url}{folder_cms_path}?action=frameset&subaction=view"
        soup = await self._get_page(url)

        files = []
        seen_urls = set()
        for link in soup.find_all("a", href=True):
            href = link["href"]
            if not any(href.lower().endswith(ext) for ext in [".xls", ".xlsx"]):
                continue
            if href in seen_urls:
                continue
            seen_urls.add(href)

            # Имя файла из URL
            filename = unquote(href.split("/")[-1])
            full_url = href if href.startswith("http") else f"{self.base_url}{href}"
            files.append((filename, full_url))

        return files

    async def _download_file(self, url: str, save_path: str) -> bool:
        """Скачивает файл по URL и сохраняет на диск."""
        try:
            async with self._session.get(url) as resp:
                if resp.status != 200:
                    logging.warning(f"HTTP {resp.status} при скачивании {url}")
                    return False
                content = await resp.read()
                os.makedirs(os.path.dirname(save_path), exist_ok=True)
                with open(save_path, "wb") as f:
                    f.write(content)
                return True
        except Exception as e:
            logging.error(f"Ошибка скачивания {url}: {e}")
            return False

    async def run(self) -> list[str]:
        """Основной процесс: логин → обход недель → факультеты → скачивание .xls.
        Возвращает список путей к скачанным файлам."""
        self.ensure_download_dir()
        downloaded_files = []

        # SSL-контекст без проверки сертификата (bb.usurt.ru использует самоподписанный)
        ssl_ctx = ssl.create_default_context()
        ssl_ctx.check_hostname = False
        ssl_ctx.verify_mode = ssl.CERT_NONE

        connector = aiohttp.TCPConnector(ssl=ssl_ctx)
        timeout = aiohttp.ClientTimeout(total=30)
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/120.0.0.0 Safari/537.36"}

        async with aiohttp.ClientSession(connector=connector, timeout=timeout, headers=headers) as session:
            self._session = session
            try:
                logging.info("--- START BB LOGIN (HTTP) ---")
                await self.login_to_bb()

                week_folders = await self._get_week_folders()
                if not week_folders:
                    logging.error("Не удалось найти папки недель.")
                    return []

                for week_name, week_path in week_folders.items():
                    logging.info(f"=== ОБРАБОТКА: {week_name.upper()} ===")

                    faculties = await self._get_faculty_folders(week_path)
                    if not faculties:
                        # Файлы могут лежать прямо в папке недели (без подпапок факультетов)
                        xls_files = await self._get_xls_links(week_path)
                        for filename, dl_url in xls_files:
                            safe_week = re.sub(r'[\\/*?:"<>|]', "_", week_name).strip()
                            save_path = os.path.join(self.download_dir, "Общее", f"{safe_week}_{filename}")
                            if await self._download_file(dl_url, save_path):
                                downloaded_files.append(save_path)
                        continue

                    for faculty_name, faculty_path in faculties.items():
                        xls_files = await self._get_xls_links(faculty_path)
                        if not xls_files:
                            continue

                        logging.info(f"  📁 {faculty_name}: {len(xls_files)} файлов")
                        for filename, dl_url in xls_files:
                            safe_week = re.sub(r'[\\/*?:"<>|]', "_", week_name).strip()
                            final_filename = f"{safe_week}_{filename}"
                            save_path = os.path.join(self.download_dir, faculty_name, final_filename)
                            if await self._download_file(dl_url, save_path):
                                downloaded_files.append(save_path)

                logging.info(f"=== ВСЕ НЕДЕЛИ ОБРАБОТАНЫ: {len(downloaded_files)} файлов скачано ===")
                return downloaded_files

            except Exception as e:
                logging.error(f"Ошибка скрапинга: {e}", exc_info=True)
                return []
            finally:
                self._session = None


class ScheduleProcessor:
    def __init__(self):
        self.db_path = DB_PATH
        self.schedules_dir = DOWNLOAD_DIR
        self.current_year = datetime.now().year

    def determine_week_type(self, filename):
        filename_lower = filename.lower()
        if 'нечетная' in filename_lower or 'нечет' in filename_lower: return 'нечетная'
        if 'четная' in filename_lower or 'чет' in filename_lower: return 'четная'
        if 'аттестация' in filename_lower or 'сессия' in filename_lower: return 'сессия'
        return 'неизвестно'

    def parse_filename_context(self, filename):
        match = re.search(r'(\d)\s*семестр.*?(\d{4})[/-](\d{4})', filename, re.IGNORECASE)
        if match:
            return int(match.group(1)), int(match.group(2)), int(match.group(3))
        return None

    def parse_date_from_cell(self, cell_content, context):
        if not isinstance(cell_content, str): return None
        
        date_match = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{2,4})', cell_content)
        if date_match:
            try:
                day = int(date_match.group(1))
                month = int(date_match.group(2))
                year_str = date_match.group(3)
                year = int(year_str) + 2000 if len(year_str) == 2 else int(year_str)
                return datetime(year, month, day).strftime('%Y-%m-%d')
            except ValueError:
                pass
                
        match = re.search(r'(\d+)\s+([а-я]+)', cell_content, re.IGNORECASE)
        if match:
            day = int(match.group(1))
            month = MONTHS_MAP.get(match.group(2).lower())
            if month:
                target_year = context.get('year', datetime.now().year)
                if 'semester' in context:
                    if context['semester'] == 1:
                        target_year = context['start_year'] if month >= 9 else context['end_year']
                    elif context['semester'] == 2:
                        target_year = context['end_year']
                else:
                    if month > 9 and datetime.now().month < 5:
                        target_year = datetime.now().year - 1
                    else:
                        target_year = datetime.now().year
                try:
                    return datetime(target_year, month, day).strftime('%Y-%m-%d')
                except ValueError:
                    return None
        return None

    def parse_lesson_cell(self, cell_content):
        if not isinstance(cell_content, str) or not cell_content.strip(): return None
        lines = [re.sub(r'^\s*-\s*', '', line).strip() for line in cell_content.split('\n') if line.strip()]
        if not lines: return None
            
        subject = lines[0]
        teacher = "Не указан"
        location = "Не указана"
        location_parts = []
        
        if len(lines) > 1:
            teacher_found = False
            for i in range(1, len(lines)):
                line = lines[i]
                if not teacher_found and line.lower() == "не указан": continue
                
                if not teacher_found:
                    is_academic = any(k in line.lower() for k in ["преподаватель", "доцент", "профессор", "ассистент", "зав. кафедрой"])
                    is_name_format = re.match(r'^[А-ЯЁ][а-яё\-]+\s+([А-ЯЁ]\.\s*[А-ЯЁ]\.|[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+)', line)
                    if is_academic or is_name_format:
                        teacher = line
                        teacher_found = True
                    else:
                        location_parts.append(line)
                else:
                    location_parts.append(line)
            location = " ".join(location_parts) if location_parts else "Не указана"
            
        subgroup_match = re.search(r'(\d\s*п/г)', cell_content, re.IGNORECASE)
        if subgroup_match:
            subject += f" ({subgroup_match.group(1).replace(' ', '')})"
            
        return {"subject": subject, "teacher": teacher, "location": location}

    def _read_rows(self, file_path: str) -> list[list[str]]:
        """Читает все строки Excel файла через xlrd (.xls) или openpyxl (.xlsx)."""
        if file_path.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows():
                rows.append([str(cell.value) if cell.value is not None else "" for cell in row])
            wb.close()
            return rows
        else:
            import xlrd
            from xlrd.biffh import XLRDError
            try:
                wb = xlrd.open_workbook(file_path)
                ws = wb.sheet_by_index(0)
                rows = []
                for rx in range(ws.nrows):
                    rows.append([str(ws.cell_value(rx, cx)) for cx in range(ws.ncols)])
                return rows
            except XLRDError as e:
                if "Excel xlsx file; not supported" in str(e):
                    # Файл называется .xls, но внутри это .xlsx
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active
                    rows = []
                    for row in ws.iter_rows():
                        rows.append([str(cell.value) if cell.value is not None else "" for cell in row])
                    wb.close()
                    return rows
                raise

    def process_single_file(self, file_path, faculty="Неизвестно", course="N/A"):
        lessons_list = []
        filename = os.path.basename(file_path)
        if not filename.endswith((".xls", ".xlsx")): return []
        
        week_type = self.determine_week_type(filename)
        if week_type == 'неизвестно': return []
        
        file_context = {}
        acad_context = self.parse_filename_context(filename)
        if acad_context:
            file_context.update({'semester': acad_context[0], 'start_year': acad_context[1], 'end_year': acad_context[2]})
        else:
            file_context['year'] = self.current_year
            
        try:
            rows = self._read_rows(file_path)
            if not rows:
                return []

            # Ищем строку заголовка (содержит "День" или "Day")
            header_row_index = -1
            for i, row in enumerate(rows):
                if 'День' in row or ('Day' in row and 'Time' in row):
                    header_row_index = i
                    break
            
            if header_row_index == -1: return []
            
            # Формируем заголовки
            headers = [h.strip() for h in rows[header_row_index]]
            
            day_col = headers.index('День') if 'День' in headers else (headers.index('Day') if 'Day' in headers else -1)
            time_col = headers.index('Часы') if 'Часы' in headers else (headers.index('Time') if 'Time' in headers else -1)
            
            if day_col == -1 or time_col == -1: return []
            
            # Индексы групп — все столбцы кроме День, Часы и пустых
            skip = {day_col, time_col}
            group_cols = [(idx, headers[idx]) for idx in range(len(headers)) 
                          if idx not in skip and headers[idx] and headers[idx] != 'nan']
            
            current_date_str = None
            current_time_slot = None
            
            for row in rows[header_row_index + 1:]:
                # Расширяем строку до нужной длины (бывают неполные строки)
                while len(row) <= max(day_col, time_col):
                    row.append("")
                
                potential_date = self.parse_date_from_cell(row[day_col], file_context)
                if potential_date:
                    current_date_str = potential_date
                    current_time_slot = None
                
                if not current_date_str: continue
                
                raw_time = row[time_col].strip() if time_col < len(row) else ""
                if raw_time and "nan" not in raw_time.lower():
                    time_slot = raw_time
                    current_time_slot = time_slot
                elif current_time_slot:
                    time_slot = current_time_slot
                else:
                    continue
                    
                for col_idx, group_name in group_cols:
                    cell_val = row[col_idx] if col_idx < len(row) else ""
                    lesson_info = self.parse_lesson_cell(cell_val)
                    if lesson_info:
                        lessons_list.append((
                            group_name.strip(), current_date_str, time_slot,
                            lesson_info['subject'], lesson_info['teacher'], lesson_info['location'],
                            week_type, faculty, course
                        ))
            return lessons_list
        except Exception as e:
            logging.error(f"Error processing {filename}: {e}")
            return []

    def run(self):
        if not os.path.exists(self.schedules_dir):
            logging.error(f"Directory {self.schedules_dir} not found.")
            return False

        conn = sqlite3.connect(self.db_path)
        all_lessons = []
        
        logging.info("Processing schedule files...")
        for dirpath, _, filenames in os.walk(self.schedules_dir):
            if dirpath == self.schedules_dir: continue
            
            relative_path = os.path.relpath(dirpath, self.schedules_dir)
            path_parts = relative_path.split(os.sep)
            faculty = path_parts[0] if path_parts else "Неизвестно"
            course_str = path_parts[1] if len(path_parts) > 1 else "Без курса"
            course = re.search(r'\d+', course_str).group(0) if re.search(r'\d+', course_str) else "N/A"
            
            for filename in filenames:
                file_path = os.path.join(dirpath, filename)
                current_file_course = course
                if current_file_course == "N/A":
                    match = re.search(r'(\d+)\s*курс', filename, re.IGNORECASE)
                    if match: current_file_course = match.group(1)
                
                lessons = self.process_single_file(file_path, faculty, current_file_course)
                all_lessons.extend(lessons)

        if not all_lessons:
            logging.warning("No lessons found.")
            conn.close()
            return False

        try:
            cursor = conn.cursor()
            cursor.execute("BEGIN TRANSACTION;")
            cursor.execute("DELETE FROM schedule;")
            cursor.executemany("""
            INSERT INTO schedule (group_name, lesson_date, time, subject, teacher, location, week_type, faculty, course)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, all_lessons)
            conn.commit()
            logging.info(f"Updated DB with {len(all_lessons)} lessons.")
            return True
        except Exception as e:
            conn.rollback()
            logging.error(f"DB Update Error: {e}")
            return False
        finally:
            conn.close()

async def run_full_sync():
    start_time = datetime.now(timezone.utc)
    logging.info(f"Начало полного цикла обновления расписания: {start_time}")
    fetcher = ScheduleFetcher()
    processor = ScheduleProcessor()
    
    details = {}
    status = "ERROR"
    try:
        xls_files = await fetcher.run()
        details["excel_files_downloaded"] = len(xls_files)
        
        # Теоретически тут можно посчитать rows_processed, 
        # но ScheduleProcessor напрямую через sqlite3 пишет в БД.
        # Для простоты считаем успешным, если не было исключений.
        if xls_files:
            processor.run()
            status = "SUCCESS"
        else:
            logging.warning("Не удалось скачать файлы расписания.")
            details["error"] = "No files downloaded"
            status = "ERROR"
            
        return status == "SUCCESS"
        
    except Exception as e:
        status = "ERROR"
        details["error"] = str(e)
        logging.exception("Ошибка при обновлении расписания")
        return False
    finally:
        end_time = datetime.now(timezone.utc)
        duration = end_time - start_time
        details["duration_seconds"] = duration.total_seconds()
        
        try:
            await save_job_log("schedule_sync", start_time, end_time, status, details)
            await cleanup_old_job_logs(days=30)
        except Exception as e_log:
            logging.error(f"Не удалось сохранить лог задачи: {e_log}")
